import pandas as pd
import json
import re
import sys
import os
from typing import Dict, List
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. FONCTIONS UTILITAIRES
# ==========================================

def clean_markdown(text: str) -> str:
    """Nettoie le markdown (gras, italique, code) pour Excel."""
    if not isinstance(text, str):
        return ""
    # Remplacement des blocs de code
    text = re.sub(r'```.*?```', '[CODE BLOCK]', text, flags=re.DOTALL)
    # Suppression du formatage inline
    text = re.sub(r'`([^`]+)`', r'\1', text)      # code inline
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text) # gras **
    text = re.sub(r'__([^_]+)__', r'\1', text)     # gras __
    text = re.sub(r'\*([^*]+)\*', r'\1', text)     # italique *
    text = re.sub(r'_([^_]+)_', r'\1', text)       # italique _
    # Suppression des liens [texte](url) -> texte
    text = re.sub(r'\[([^\]]+)\]\(([^)]+)\)', r'\1', text)
    # Suppression des titres #
    text = re.sub(r'^#+\s+', '', text, flags=re.MULTILINE)
    return text.strip()

def format_evidence_list(evidence_list: List[str]) -> str:
    """Transforme une liste de preuves en texte à puces."""
    if not evidence_list:
        return ""
    # Nettoyage de chaque preuve
    cleaned_items = [clean_markdown(str(e)) for e in evidence_list if e]
    return "\n".join([f"• {item}" for item in cleaned_items])

def get_severity_color(severity: str) -> str:
    """Retourne le code HEX pour la couleur de fond selon la sévérité."""
    colors = {
        "CRITICAL": "DC2626", # Rouge vif
        "HIGH": "EA580C",     # Orange
        "MEDIUM": "CA8A04",   # Jaune moutarde
        "LOW": "16A34A",      # Vert
        "UNKNOWN": "6B7280"   # Gris
    }
    return colors.get(str(severity).upper(), "FFFFFF")

# ==========================================
# 2. GÉNÉRATEUR EXCEL
# ==========================================

def generate_excel_report(judgment_data: Dict, output_filename: str) -> str:
    """Génère le fichier Excel complet."""
    
    # --- A. Extraction des données ---
    # Gestion souple de la structure JSON (avec ou sans clé racine 'judgment')
    if "judgment" in judgment_data and "executive_summary" not in judgment_data:
        root = judgment_data["judgment"]
    else:
        root = judgment_data

    summary = root.get("executive_summary", {})
    confirmed_issues = root.get("confirmed_issues", [])
    dismissed_issues = root.get("dismissed_issues", [])
    needs_verification = root.get("needs_verification", [])

    # --- B. Préparation du DASHBOARD ---
    metrics_data = {
        "Métrique": [
            "Total Issues", "Critical", "High", "Medium", "Low", 
            "Greenwashing Count", "Risk Level", "To Verify"
        ],
        "Valeur": [
            summary.get('total_confirmed_issues', 0),
            summary.get('critical_issues', 0),
            summary.get('high_issues', 0),
            summary.get('medium_issues', 0),
            summary.get('low_issues', 0),
            summary.get("greenwashing_count", 0),
            summary.get("greenwashing_risk_level", "N/A"),
            len(needs_verification)
        ]
    }
    df_metrics = pd.DataFrame(metrics_data)

    # --- C. Préparation des CONFIRMED ISSUES (Multi-Analystes) ---
    issues_rows = []
    for issue in confirmed_issues:
        
        # 1. Gestion des contributions multiples
        analyst_contribs = issue.get("analyst_contributions", [])
        
        # Concaténation des descriptions : [NOM] Description
        desc_parts = []
        all_evidence = []
        
        if analyst_contribs:
            for contrib in analyst_contribs:
                name = contrib.get("analyst", "Analyste ?")
                desc = clean_markdown(contrib.get("description", "Pas de description"))
                desc_parts.append(f"[{name.upper()}]\n{desc}")
                all_evidence.extend(contrib.get("evidence", []))
        else:
            # Fallback si pas de contributions détaillées
            desc_parts.append(clean_markdown(issue.get("description", "")))
            all_evidence.extend(issue.get("evidence", []))

        # Jointure avec séparateur visuel
        full_description = "\n\n" + ("-"*20) + "\n\n".join(desc_parts)
        
        # Dé-doublonnage des preuves
        unique_evidence = sorted(list(set(all_evidence)), key=len, reverse=True)
        formatted_evidence = format_evidence_list(unique_evidence)

        # Création de la ligne
        issues_rows.append({
            "ID": issue.get('final_id', 'N/A'),
            "Title": clean_markdown(issue.get('title', 'Untitled')),
            "Severity": issue.get("final_severity", issue.get("severity", "MEDIUM")),
            "Type": issue.get("type", "N/A"),
            "Is GW": "YES" if issue.get("type") == "GREENWASHING" else "NO",
            "Nb Analysts": len(analyst_contribs),
            "Page Refs": ", ".join(map(str, issue.get("all_page_references", []))),
            "Confidence": issue.get("consensus", {}).get("confidence", "N/A"),
            "Description (Consolidated)": full_description.strip(),
            "Evidence (All)": formatted_evidence
        })
        
    df_issues = pd.DataFrame(issues_rows)

    # --- D. Préparation des autres onglets ---
    # To Verify
    verify_rows = []
    for v in needs_verification:
        verify_rows.append({
            "ID": v.get("issue_id", "N/A"),
            "Title": clean_markdown(v.get("title", "")),
            "Reason": clean_markdown(v.get("verification_reason", "")),
            "Action Required": clean_markdown(v.get("what_to_check", ""))
        })
    df_verify = pd.DataFrame(verify_rows)

    # Dismissed
    dismissed_rows = []
    for d in dismissed_issues:
        dismissed_rows.append({
            "ID": d.get("issue_id", "N/A"),
            "Title": clean_markdown(d.get("title", "")),
            "Reason": clean_markdown(d.get("dismissal_reason", ""))
        })
    df_dismissed = pd.DataFrame(dismissed_rows)

    # --- E. Écriture et Formatage Excel ---
    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            
            # 1. Écriture des données
            df_metrics.to_excel(writer, sheet_name='Dashboard', index=False, startrow=1, startcol=1)
            if not df_issues.empty: 
                df_issues.to_excel(writer, sheet_name='Confirmed Issues', index=False)
            if not df_verify.empty: 
                df_verify.to_excel(writer, sheet_name='To Verify', index=False)
            if not df_dismissed.empty: 
                df_dismissed.to_excel(writer, sheet_name='Dismissed', index=False)

            # 2. Formatage
            workbook = writer.book
            
            # --- Style Global (Headers) ---
            header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for ws in workbook.worksheets:
                # Appliquer style aux en-têtes (Ligne 1 pour la plupart, Ligne 2 pour Dashboard)
                start_row = 2 if ws.title == 'Dashboard' else 1
                for cell in ws[start_row]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # --- Style Spécifique : Confirmed Issues ---
            if 'Confirmed Issues' in workbook.sheetnames:
                ws = workbook['Confirmed Issues']
                
                # Index des colonnes (1-based)
                col_sev = 3   # Severity
                col_gw = 5    # Is GW
                col_desc = 9  # Description
                col_evid = 10 # Evidence

                # Largeurs de colonnes optimisées
                ws.column_dimensions['B'].width = 30  # Title
                ws.column_dimensions['I'].width = 70  # Description (Large !)
                ws.column_dimensions['J'].width = 60  # Evidence (Large !)
                
                # Itération sur les lignes de données
                for row in range(2, ws.max_row + 1):
                    # Alignement haut et retour à la ligne pour tout le monde
                    for cell in ws[row]:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    # Couleur Sévérité
                    sev_cell = ws.cell(row=row, column=col_sev)
                    if sev_cell.value:
                        c = get_severity_color(str(sev_cell.value))
                        sev_cell.fill = PatternFill(start_color=c, end_color=c, fill_type="solid")
                        sev_cell.font = Font(color="FFFFFF", bold=True)
                        sev_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Alerte Greenwashing
                    gw_cell = ws.cell(row=row, column=col_gw)
                    if gw_cell.value == "YES":
                        gw_cell.font = Font(color="DC2626", bold=True)
                        gw_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        gw_cell.alignment = Alignment(horizontal='center', vertical='center')

            # --- Style Dashboard ---
            ws_dash = workbook['Dashboard']
            ws_dash['B1'] = "CSRD REPORT DASHBOARD"
            ws_dash['B1'].font = Font(size=14, bold=True, color="2563EB")
            ws_dash.column_dimensions['B'].width = 25
            ws_dash.column_dimensions['C'].width = 15

        return output_filename

    except Exception as e:
        return f"Erreur critique lors de l'écriture Excel: {str(e)}"

# ==========================================
# 3. EXÉCUTION DU SCRIPT
# ==========================================

if __name__ == "__main__":
    # Vérification des arguments
    if len(sys.argv) < 2:
        print("Usage: python json_to_excel.py <fichier_entree.json> [fichier_sortie.xlsx]")
        sys.exit(1)

    input_file = sys.argv[1]
    
    # Gestion intelligente du nom de fichier de sortie
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
        # Correction automatique de l'extension
        if output_file.lower().endswith('.json'):
            output_file = os.path.splitext(output_file)[0] + ".xlsx"
            print(f"⚠️  Extension corrigée : sortie vers '{output_file}'")
        elif not output_file.lower().endswith('.xlsx'):
            output_file += ".xlsx"
    else:
        # Par défaut : nom du json avec extension .xlsx
        base_name = os.path.splitext(input_file)[0]
        output_file = base_name + ".xlsx"

    print(f"📖 Lecture du fichier : {input_file}")
    
    try:
        # Lecture du JSON
        with open(input_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        print(f"⚙️  Conversion en cours...")
        result = generate_excel_report(data, output_file)
        
        if "Erreur" in str(result) and not result.endswith('.xlsx'):
             print(f"❌ {result}")
        else:
             print(f"✅ Succès ! Fichier généré : {result}")
             print(f"   (Les descriptions de {len(data.get('confirmed_issues', []))} issues ont été consolidées)")
            
    except FileNotFoundError:
        print(f"❌ Erreur : Le fichier '{input_file}' est introuvable.")
    except json.JSONDecodeError:
        print(f"❌ Erreur : Le fichier '{input_file}' n'est pas un JSON valide.")
    except Exception as e:
        print(f"❌ Une erreur inattendue est survenue : {e}")